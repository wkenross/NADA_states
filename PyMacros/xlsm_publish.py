import os
import openpyxl
import win32com.client as win32
from openpyxl import load_workbook
from win32com.client import Dispatch, GetActiveObject, GetObject

# Check if Excel is installed
try:
    xl = GetActiveObject("Excel.Application")
except:
    # Open Excel if it is not already running
    xl = Dispatch("Excel.Application")

# Check if Excel is visible
if not xl.Visible: xl.Visible = True

# Set the path to the original macro enabled workbook
original_path = r'C:\projects\PyMacros\vba_macros.xlsm'

# Open the original macro enabled workbook
wb = load_workbook(filename=original_path, keep_vba=True)

# Check if the workbook has VBA modules
if hasattr(wb, 'vba_project'):
    # Iterate over all sheets in the workbook
    for sheet in wb.sheetnames:
        # Check if the sheet is the workbook object
        if sheet == 'ThisWorkbook': 
            continue
        else:
            # Iterate over all modules attached to the workbook
            for module in wb.vba_project.modules:
                # Check if the module name matches the name to be removed
                if module.name == 'Module1':
                    # Remove the module from the workbook
                    wb.vba_project.remove_module(module)

# Iterate over all sheets in the workbook
for sheet in wb.sheetnames:
    # Check if the sheet has VBA modules
    if hasattr(wb[sheet], '_vba_controls'):
        # Iterate over all modules attached to the sheet
        for name in wb[sheet]._vba_controls:
            # Remove the module from the sheet
            wb[sheet]._vba_controls[name] = None

# Set the save path for the new xlsx file
save_path = os.path.splitext(original_path)[0] + '.xlsx'

# Save a copy of the workbook as a xlsx file
# wb.SaveAs(save_path, FileFormat=51)

# Remove any data links in the XLSX file
new_wb = load_workbook(save_path)
for link in new_wb._external_links:
    new_wb._external_links.remove(link)
    new_wb.save(save_path)

# Activate the orginal workbook and new workbook
wb.active = True
new_wb.active = True



# Open the original macro enabled workbook and activate it
# xl = win32.gencache.EnsureDispatch('Excel.Application')
# xl.Visible = True
# wb = xl.Workbooks.Open(original_path)
# wb.Activate()

# Remove any modules in the workbook
# for module in wb.VBProject.VBComponents:
    # wb.VBProject.VBComponents.Remove(module)

# Remove all VBA modules
# for name in wb.vba_modules:
    # del wb.vba_modules[name]

# Remove any macros or modules behind a sheet
# for sheet in wb.Worksheets:
#     if sheet.CodeName != 'ThisWorkbook':
#         sheet.CodeName = sheet.Name
#         sheet.DrawingObjects.Delete()


# Close the workbook
# wb.close()
# xl.Quit()
